import os
import re
import json
import pathlib
import openpyxl
from openpyxl import Workbook

def check_and_set_data_file(data_file_path="data.xlsx"):
    data_file = pathlib.Path(data_file_path)
    
    if not data_file.exists():
        workbook = Workbook()
        sheet = workbook.active

        sheet["A1"] = "Student No."
        sheet["B1"] = "Python"
        sheet["C1"] = "SQL"
        sheet["D1"] = "HTML"
        sheet["E1"] = "Javascript"
        sheet["F1"] = "ASP.Net"
        sheet["G1"] = "C#"

        workbook.save(data_file_path)

def read_data_file_and_set_database_and_subjects(data_file_path="data.xlsx"):
    global database, subjects

    workbook = openpyxl.load_workbook(data_file_path)
    sheet = workbook.active

    for i, r in enumerate(sheet.iter_rows(min_row=1)):
        if i == 0:
            for j in range(1, len(r)):
                subjects.append(r[j].value)
        else:
            database[r[0].value] = []
            for j in range(1, len(subjects) + 1):
                database[r[0].value].append(r[j].value)

def get_welcome_console(institute_name="ABC Computer Training", diploma_name="Computer Literacy"):
    data = f'''
            WELCOME TO ACADEMY OF {institute_name.upper()} 
            DIPLOMA PROGRAMME: {diploma_name.upper()}\n
                        [MAIN MENU]

                    1. Enter Student Marks
                    2. Display Student Marks
                    3. Delete Student Marks
                    4. Display Results of a Given Student
                    5. Display Results of all Students
                    6. Exit
            '''
    return data

def get_grade_for_marks(marks):
    if marks < 0:
        return "N/A"
    if marks >= 70:
        return 'A'
    if marks >= 50:
        return 'B'
    if marks >= 30:
        return 'C'
    return 'D'

def get_results_of_students(ids):
    for id in ids:
        marks = database[id]
        grades = [get_grade_for_marks(m) for m in marks]
        total = sum(marks)
        avg = total / len(subjects)
        record = {"Student No.": id}

        for i, s in enumerate(subjects):
            record[s] = {"Marks": marks[i], "Grade": grades[i]}

        record["total"] = total
        record["average"] = round(avg, 2)
        record["Final Result"] = "PASS"

        if avg < 50:
            record["Final Result"] = "FAIL"
        record = json.dumps(record, indent=11)
        record = re.sub('"', '', record)

        yield record

def update_and_save_data_file(data_file_path="data.xlsx"):
    global database, subjects
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet["A1"] = "Student No."
    sheet["B1"] = "Python"
    sheet["C1"] = "SQL"
    sheet["D1"] = "HTML"
    sheet["E1"] = "Javascript"
    sheet["F1"] = "ASP.Net"
    sheet["G1"] = "C#"

    for r, id in enumerate(list(database.keys())):
        r += 2
        id_cell = sheet.cell(row=r, column=1)
        id_cell.value = id

        for c in range(2, len(subjects) + 2):
            cell = sheet.cell(row=r, column=c)
            cell.value = database[id][c - 2]

    workbook.save(data_file_path)

database = dict()
subjects = []
data_file_path = "data file.xlsx"

check_and_set_data_file(data_file_path)
read_data_file_and_set_database_and_subjects(data_file_path)
state = ""

while state.lower() != "q":
    os.system("cls")
    print(get_welcome_console())
    
    choice = input("\t    User Option: ")

    if choice == '1':
        os.system("cls")
        print("\t   [ENTER STUDENT MARKS]\n")
        
        id = input("\t   Student ID: ")

        if id[0].lower() != 's':
            id = 'S' + id

        database[id] = []

        for s in subjects:
            marks = int(input("\t   Subject-" + s + " Marks: "))
            database[id].append(marks)
        
        print("\n\t   Successfully Recorded the New Student Mark Entry")
    
    elif choice == '2':
        os.system("cls")
        print("\t   [DISPLAY STUDENT MARKS]\n")
        
        for k in database.keys():
            record = dict(zip(subjects, database[k]))
            record["Student No."] = k
            record = json.dumps(record, indent=11)
            record = re.sub('"', '', record)

            print(record)
        
        print("\n\t   Successfully Retrieved Student Mark Entries")

    elif choice == '3':
        os.system("cls")
        print("\t   [DELETE STUDENT MARKS]\n")

        id = input("\t   Student ID: ")

        if id in database:
            database.pop(id)
            print(f"\n\t   Successfully Deleted Student-{id} Entry")
        else:
            print(f"\n\t   Student-{id} Entry Is Not Available")

    elif choice == '4':
        os.system("cls")
        print("\t   [DISPLAY RESULTS OF A GIVEN STUDENT]\n")

        id = input("\t   Student ID: ")
        
        if id in database:
            record = get_results_of_students([id])

            for r in record:
                print(r)
            print(f"\n\t   Successfully Retrieved Student-{id} Results")
        else:
            print(f"\n\t   Student-{id} Results Are Not Available")

    elif choice == '5':
        os.system("cls")
        print("\t   [DISPLAY RESULTS OF ALL STUDENTS]\n")
        
        record = get_results_of_students(list(database.keys()))

        for r in record:
            print(r)

    elif choice == '6':
        os.system("cls")
        print("\t   [EXIT]\n")

        confirmation = input("\t   Are You Sure Do You Really Want To Quit the App? (Y/N) ")

        if confirmation.lower() == 'y':
            break
    else:
        print("\t   Invalid Option Entered")

    state = input("\n\t   Press Enter to Continue or q/Q to Quit... ")

update_and_save_data_file(data_file_path)

print("\n\t   Application Terminated. Checkout the Excel Data File for Further Details")
print("\t   Thank You :)")
